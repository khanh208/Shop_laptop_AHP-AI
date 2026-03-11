import csv
import json
import os
import re
import uuid
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from sqlalchemy import text


DEFAULT_SHEET_NAME = "Laptop_Data"


def _fetch_one(conn, sql, params=None):
    return conn.execute(text(sql), params or {}).mappings().first()


def _fetch_all(conn, sql, params=None):
    return conn.execute(text(sql), params or {}).mappings().all()


def _normalize_header(value):
    if value is None:
        return ""
    value = str(value).strip().lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return value.strip("_")


def _to_str(value):
    if value is None:
        return None
    value = str(value).strip()
    return value if value != "" else None


def _to_float(value):
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip()
    if s == "":
        return None

    s = s.replace(",", "")
    try:
        return float(s)
    except Exception:
        return None


def _to_int(value):
    number = _to_float(value)
    if number is None:
        return None
    return int(round(number))


def _first_non_null(row, keys):
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return None


def _compose_name(company, model):
    company = _to_str(company)
    model = _to_str(model)
    if company and model:
        return f"{company} {model}"
    return company or model


def _is_blank_row(row):
    return all(v in (None, "", "None") for v in row.values())


def _read_excel_rows(file_path, sheet_name):
    wb = load_workbook(file_path, data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Không tìm thấy sheet '{sheet_name}'. "
            f"Các sheet hiện có: {', '.join(wb.sheetnames)}"
        )

    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    normalized_headers = [_normalize_header(h) for h in headers]

    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {}
        for idx, value in enumerate(values):
            key = normalized_headers[idx] if idx < len(normalized_headers) else f"col_{idx+1}"
            row[key] = value

        if _is_blank_row(row):
            continue

        rows.append(row)

    return rows


def _read_csv_rows(file_path):
    rows = []

    encodings = ["utf-8-sig", "utf-8", "cp1258", "latin-1"]
    last_error = None

    for encoding in encodings:
        try:
            with open(file_path, "r", encoding=encoding, newline="") as f:
                reader = csv.DictReader(f)
                normalized_fieldnames = [_normalize_header(h) for h in reader.fieldnames or []]

                for raw_row in reader:
                    row = {}
                    for i, raw_key in enumerate(reader.fieldnames or []):
                        key = normalized_fieldnames[i]
                        row[key] = raw_row.get(raw_key)

                    if _is_blank_row(row):
                        continue

                    rows.append(row)
            return rows
        except Exception as e:
            last_error = e

    raise ValueError(f"Không đọc được file CSV: {last_error}")


def _map_staging_row(row, source_file_name, batch_id):
    company = _to_str(_first_non_null(row, ["company"]))
    full_name = _to_str(_first_non_null(row, ["full_name"]))
    product = _to_str(_first_non_null(row, ["product"]))

    cpu_company = _to_str(_first_non_null(row, ["cpu_company"]))
    cpu_type = _to_str(_first_non_null(row, ["cpu_type"]))
    gpu_company = _to_str(_first_non_null(row, ["gpu_company"]))
    gpu_type = _to_str(_first_non_null(row, ["gpu_type"]))

    battery_hours = _to_float(_first_non_null(row, ["battery_hrs", "battery_fill"]))
    durability_score = _to_float(_first_non_null(row, ["durability_1_10", "durability_fill"]))
    upgradeability_score = _to_float(_first_non_null(row, ["upgradability_1_10", "upgradability_fill"]))

    laptop_name = full_name or _compose_name(company, product) or "Unknown Laptop"
    model_code = product or ""

    cpu_name = _compose_name(cpu_company, cpu_type) or cpu_type or cpu_company or "Unknown CPU"
    gpu_name = _compose_name(gpu_company, gpu_type) or gpu_type or gpu_company

    description_parts = []
    for key in ["typename", "condition", "memory", "screensresolution", "screenresolution", "opsys"]:
        value = _to_str(row.get(key))
        if value:
            description_parts.append(value)

    description = " | ".join(description_parts) if description_parts else None

    return {
        "import_batch": batch_id,
        "source_file_name": source_file_name,

        "brand_name": company,
        "laptop_name": laptop_name,
        "model_code": model_code,

        "cpu_name": cpu_name,
        "cpu_benchmark_score": _to_float(_first_non_null(row, ["cpu_score_1_10"])),
        "gpu_name": gpu_name,
        "gpu_benchmark_score": _to_float(_first_non_null(row, ["gpu_score_1_10"])),

        "ram_gb": _to_int(_first_non_null(row, ["ram_gb"])),
        "ssd_gb": _to_int(_first_non_null(row, ["ssd_gb"])),
        "screen_size_inch": _to_float(_first_non_null(row, ["inches"])),
        "screen_resolution": _to_str(_first_non_null(row, ["screenresolution", "screensresolution"])),
        "refresh_rate_hz": None,
        "weight_kg": _to_float(_first_non_null(row, ["weight_kg"])),
        "battery_hours": battery_hours,

        "durability_score": durability_score,
        "upgradeability_score": upgradeability_score,

        "price": _to_float(_first_non_null(row, ["price_vnd"])),
        "image_url": None,
        "product_url": None,
        "description": description,

        "Norm_CPU": _to_float(_first_non_null(row, ["norm_cpu"])),
"Norm_RAM": _to_float(_first_non_null(row, ["norm_ram"])),
"Norm_GPU": _to_float(_first_non_null(row, ["norm_gpu"])),
"Norm_Screen": _to_float(_first_non_null(row, ["norm_screen"])),
"Norm_Weight": _to_float(_first_non_null(row, ["norm_weight"])),
"Norm_Battery": _to_float(_first_non_null(row, ["norm_battery"])),
"Norm_Durability": _to_float(_first_non_null(row, ["norm_durability"])),
"Norm_Upgrade": _to_float(_first_non_null(row, ["norm_upgrade"])),

"price_vnd_col": _to_float(_first_non_null(row, ["price_vnd"])),
"ahp_score_col": _to_float(_first_non_null(row, ["ahp_score"])),

        "raw_row": json.dumps(row, ensure_ascii=False, default=str),
    }


def stage_laptop_file(conn, file_path, original_filename, sheet_name=DEFAULT_SHEET_NAME, replace_staging=False):
    ext = Path(original_filename).suffix.lower()

    if replace_staging:
        conn.execute(text("DELETE FROM stg_laptop_data"))

    if ext in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        raw_rows = _read_excel_rows(file_path, sheet_name)
    elif ext == ".csv":
        raw_rows = _read_csv_rows(file_path)
    else:
        raise ValueError("Chỉ hỗ trợ file Excel (.xlsx) hoặc CSV (.csv)")

    if not raw_rows:
        raise ValueError("File không có dữ liệu để import")

    batch_id = f"batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"

    staged_rows = []
    skipped_rows = 0

    for row in raw_rows:
        mapped = _map_staging_row(row, original_filename, batch_id)

        # bỏ qua row quá rỗng
        if not mapped["brand_name"] and mapped["laptop_name"] == "Unknown Laptop":
            skipped_rows += 1
            continue

        staged_rows.append(mapped)

    if not staged_rows:
        raise ValueError("Không map được dữ liệu hợp lệ từ file")

    conn.execute(text("""
        INSERT INTO stg_laptop_data (
            import_batch,
            source_file_name,

            brand_name,
            laptop_name,
            model_code,

            cpu_name,
            cpu_benchmark_score,
            gpu_name,
            gpu_benchmark_score,

            ram_gb,
            ssd_gb,
            screen_size_inch,
            screen_resolution,
            refresh_rate_hz,
            weight_kg,
            battery_hours,

            durability_score,
            upgradeability_score,

            price,
            image_url,
            product_url,
            description,

            "Norm_CPU",
"Norm_RAM",
"Norm_GPU",
"Norm_Screen",
"Norm_Weight",
"Norm_Battery",
"Norm_Durability",
"Norm_Upgrade",
"Price (VND)",
"AHP Score",

raw_row
)
VALUES (
:import_batch,
:source_file_name,

:brand_name,
:laptop_name,
:model_code,

:cpu_name,
:cpu_benchmark_score,
:gpu_name,
:gpu_benchmark_score,

:ram_gb,
:ssd_gb,
:screen_size_inch,
:screen_resolution,
:refresh_rate_hz,
:weight_kg,
:battery_hours,

:durability_score,
:upgradeability_score,

:price,
:image_url,
:product_url,
:description,

:Norm_CPU,
:Norm_RAM,
:Norm_GPU,
:Norm_Screen,
:Norm_Weight,
:Norm_Battery,
:Norm_Durability,
:Norm_Upgrade,
:price_vnd_col,
:ahp_score_col,

CAST(:raw_row AS JSONB)
)
    """), staged_rows)

    return {
        "batchId": batch_id,
        "sourceFileName": original_filename,
        "sheetName": sheet_name if ext != ".csv" else None,
        "rowCount": len(staged_rows),
        "skippedRowCount": skipped_rows,
        "replaceStaging": replace_staging,
    }


def preview_staging(conn, batch_id=None, limit=50):
    if limit < 1:
        limit = 1
    if limit > 200:
        limit = 200

    if not batch_id:
        latest_batch = _fetch_one(conn, """
            SELECT import_batch
            FROM stg_laptop_data
            ORDER BY imported_at DESC, id DESC
            LIMIT 1
        """)
        batch_id = latest_batch["import_batch"] if latest_batch else None

    if not batch_id:
        return {
            "batchId": None,
            "totalRows": 0,
            "rows": []
        }

    total = _fetch_one(conn, """
        SELECT COUNT(*) AS cnt
        FROM stg_laptop_data
        WHERE import_batch = :batch_id
    """, {"batch_id": batch_id})

    rows = _fetch_all(conn, """
        SELECT
            id,
            import_batch,
            source_file_name,
            brand_name,
            laptop_name,
            model_code,
            cpu_name,
            cpu_benchmark_score,
            gpu_name,
            gpu_benchmark_score,
            ram_gb,
            ssd_gb,
            screen_size_inch,
            weight_kg,
            battery_hours,
            durability_score,
            upgradeability_score,
            COALESCE("Price (VND)", price) AS price_vnd,
            "Norm_CPU",
            "Norm_RAM",
            "Norm_GPU",
            "Norm_Screen",
            "Norm_Weight",
            "Norm_Battery",
            "Norm_Durability",
            "Norm_Upgrade",
            "AHP Score"
        FROM stg_laptop_data
        WHERE import_batch = :batch_id
        ORDER BY id
        LIMIT :limit
    """, {
        "batch_id": batch_id,
        "limit": limit
    })

    return {
        "batchId": batch_id,
        "totalRows": int(total["cnt"] or 0),
        "rows": [dict(r) for r in rows]
    }


def commit_staging(conn):
    staging_count = _fetch_one(conn, "SELECT COUNT(*) AS cnt FROM stg_laptop_data")
    if not staging_count or int(staging_count["cnt"] or 0) == 0:
        raise ValueError("Staging đang trống, chưa có dữ liệu để commit")

    before = {
        "brands": int((_fetch_one(conn, "SELECT COUNT(*) AS cnt FROM brands") or {}).get("cnt", 0)),
        "laptops": int((_fetch_one(conn, "SELECT COUNT(*) AS cnt FROM laptops") or {}).get("cnt", 0)),
        "features": int((_fetch_one(conn, "SELECT COUNT(*) AS cnt FROM laptop_ml_features") or {}).get("cnt", 0)),
        "labels": int((_fetch_one(conn, "SELECT COUNT(*) AS cnt FROM laptop_training_labels") or {}).get("cnt", 0)),
    }

    conn.execute(text("CALL sp_import_laptop_data()"))

    after = {
        "brands": int((_fetch_one(conn, "SELECT COUNT(*) AS cnt FROM brands") or {}).get("cnt", 0)),
        "laptops": int((_fetch_one(conn, "SELECT COUNT(*) AS cnt FROM laptops") or {}).get("cnt", 0)),
        "features": int((_fetch_one(conn, "SELECT COUNT(*) AS cnt FROM laptop_ml_features") or {}).get("cnt", 0)),
        "labels": int((_fetch_one(conn, "SELECT COUNT(*) AS cnt FROM laptop_training_labels") or {}).get("cnt", 0)),
    }

    return {
        "message": "Đã commit staging sang bảng chính",
        "stagingRows": int(staging_count["cnt"] or 0),
        "before": before,
        "after": after,
    }


def clear_staging(conn, batch_id=None):
    if batch_id:
        count_row = _fetch_one(conn, """
            SELECT COUNT(*) AS cnt
            FROM stg_laptop_data
            WHERE import_batch = :batch_id
        """, {"batch_id": batch_id})
        deleted_count = int(count_row["cnt"] or 0)

        conn.execute(text("""
            DELETE FROM stg_laptop_data
            WHERE import_batch = :batch_id
        """), {"batch_id": batch_id})

        return {
            "message": f"Đã xóa staging của batch {batch_id}",
            "deletedCount": deleted_count,
            "batchId": batch_id,
        }

    count_row = _fetch_one(conn, "SELECT COUNT(*) AS cnt FROM stg_laptop_data")
    deleted_count = int(count_row["cnt"] or 0)

    conn.execute(text("DELETE FROM stg_laptop_data"))

    return {
        "message": "Đã xóa toàn bộ staging",
        "deletedCount": deleted_count,
        "batchId": None,
    }